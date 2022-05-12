from distutils.command.upload import upload
import requests as requests
import openpyxl
from pathlib import Path
import json


from envs import env

constants = env.env()


def get_sheet():
    upload_file = constants.FILE
    xls_file = Path('./', upload_file)
    wb_obj = openpyxl.load_workbook(xls_file)
    sheet = wb_obj.active
    return sheet
    

def run_sheet(sheet):
    repeat_devices = 0
    created_devices = 0

    if sheet.max_row > 0:
        indice = 1
        while indice <= sheet.max_row:
            cell_name_B = 'B' + str(indice)
            cell_name_D = 'D' + str(indice)
            if confirm_device(sheet[cell_name_D].value):
                print(f"Device {sheet[cell_name_B].value} already exist")
                repeat_devices = repeat_devices + 1
            else :
                create_device(sheet, cell_name_B, cell_name_D)
                created_devices = created_devices + 1     
            indice = indice + 1

        print(f"Process finish => devices created: {created_devices} | devices repeats: {repeat_devices} " )


def confirm_device(ip):

    data = {
        "name": ip,
        "tags": []
    }

    url = f"{constants.URL_BASE}{constants.VERIFY_DEVICE}"
    response = requests.post(url=url, json=data, headers={"x-api": constants.DEVICE_TOKEN}, verify=False)
    try: 
        device = response.json()
    except: 
        pass

    if len(device["data"]) > 0:
        return True
    else:
        return False       


def create_device(sheet, hostname, ip):
    data_device = {
        "host" : sheet[hostname].value,
        "id_confparameters": "5ffc8910ffaf1d68559f10b6", # Default config
        "ip": sheet[ip].value,
        "model":"cisco",
        "name" : sheet[hostname].value,
        "password_template" : 1 ,
        "status" : "active",
        "type" : "cisco_ios",
        "vendor" : "Cisco",
        "tag": '{}'
    }

    url = f"{constants.URL_BASE}{constants.CREATE_DEVICE}"
    #data_device = json.dumps(data_device, indent= 4)

    response = requests.post(url=url, json= data_device, headers={"x-api": constants.DEVICE_TOKEN}, verify=False)
    try:
        res = response.json()
        print(res, sheet[hostname].value)
    except:
        print("Fail to create")


def main():
    sheet = get_sheet()
    run_sheet(sheet)


if __name__ == '__main__':
    main()


