import openpyxl
from pathlib import Path


def create_file():
    wb = openpyxl.Workbook()
    note = wb.active
    note.title = "config"
    note.append(("HOSTNAME", "IP", "CONFIGURATION", "ERR"))
    wb.save("Config_retriever_by_devices.xlsx")


def modify_file(upload_file, data):
    xls_file = Path('./', upload_file)
    wb = openpyxl.load_workbook(xls_file)
    sheet = wb.active
    sheet.append(data)
    wb.save(upload_file)


def create_file_dynamic(filename, colum, title):
    wb = openpyxl.Workbook()
    note = wb.active
    note.title = title
    note.append(colum)
    wb.save(filename)
